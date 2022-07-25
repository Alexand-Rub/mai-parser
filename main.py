import sqlite3 as sq
from parsing import *
from openpyxl import Workbook
from os import path, mkdir


def get_articles():
    db = sq.connect('parser.db3')
    cur = db.cursor()

    sources = cur.execute("SELECT * FROM sources").fetchall()
    print('\nДоступные источники:')
    for source in sources:
        print('{num}. {name}'.format(num=source[0], name=source[1]))
    source_num = int(input('\nВыберете номер источника: '))

    sections = cur.execute("SELECT * FROM sections WHERE source_id=source_id".format(source_id=source_num)).fetchall()
    print('\nДоступные разделы:')
    for section in sections:
        print('{num}. {name}'.format(num=sections.index(section)+1, name=section[1]))
    section_num = int(input('\nВыберете номер раздела: '))
    section_link = cur.execute("SELECT link FROM sections WHERE id={id}".format(
        id=sections[section_num-1][0]
    )).fetchall()[0][0]

    match source_num:
        case 1: return get_microwave_journal(link=section_link)


def exel_maker(export_date):
    book = Workbook()
    book.create_sheet('Статьи')
    del book['Sheet']
    sheet = book['Статьи']
    row, column = 1, 1
    for article in export_date['articles']:
        for title in article:
            match row:
                case 1:
                    sheet.cell(row=row, column=column).value = title
                    sheet.cell(row=row+1, column=column).value = article[title]
                case _: sheet.cell(row=row, column=column).value = article[title]
            column += 1
        column = 1
        match row:
            case 1: row += 2
            case _: row += 1
    book.create_sheet('Теги')
    sheet = book['Теги']
    sheet.cell(row=1, column=1).value = 'Тег'
    sheet.cell(row=1, column=2).value = 'Кол-во'
    row = 1
    for tag in sorted(export_date['tags'], key=export_date['tags'].get, reverse=True):
        sheet.cell(row=row, column=1).value = tag
        sheet.cell(row=row, column=2).value = export_date['tags'][tag]
        row += 1

    if not(path.exists('tables')):
        mkdir('tables')
    table_name = 'tables/' + input('\nВведите имя таблицы: ') + '.xlsx'
    if path.exists(table_name):
        name_flag = True
    else:
        book.save(table_name)
        name_flag = False
    while name_flag:
        answer = input('\nТакой файл уже существует!\n'
                        '1. Переименовать\n'
                        '2. Перезаписать\n'
                        '3. Закончить\n'
                        '\nВведите номер команды: ')
        match answer:
            case '1':
                table_name = 'tables/' + input('Введите имя таблицы: ') + '.xlsx'
                print('\nТаблица сохранена')
            case '2':
                book.save(table_name)
                name_flag = False
                print('\nТаблица сохранена')
            case '3': name_flag = False
            case _: print('\nТакой команды нет')
    book.close()


if __name__ == '__main__':
    while True:
        print('Доступные команды:\n1. Скачать статьи \n0. Закрыть программу')
        command = input('\nВыберете номер команды: ')
        # command = 1
        match command:
            case '0': break
            case '1':
                exel_maker(get_articles())
                break
            case _: print('\nНеккоректная команда!\n')

    print('\nКонец программы. Хорошего дня!')
