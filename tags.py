import openpyxl

if __name__ == '__main__':
    book = openpyxl.load_workbook('Aerospace.xlsx')
    sheet = book.active
    tags_count = dict()
    for i in range(2, 1000):
        if not sheet.cell(row=i, column=2).value is None:
            tags = sheet.cell(row=i, column=2).value.split(', ')
            for tag in tags:
                if tag in tags_count.keys():
                    tags_count[tag] += 1
                else:
                    tags_count[tag] = 1

    book.create_sheet('tags')
    sheet = book['tags']
    line = 1
    for k in sorted(tags_count, key=tags_count.get, reverse=True):
        print(f'{k}: {tags_count[k]}')
        sheet.cell(row=line, column=1).value = k
        sheet.cell(row=line, column=2).value = tags_count[k]
        line += 1
    book.save('Aerospace.xlsx')
    book.close()








